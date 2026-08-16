"""Tests for the one-click Excel workbook export (v0.9.0)."""
from __future__ import annotations

import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from demo_seed import DEMO_OWNER_EMAIL, DEMO_OWNER_PASSWORD, seed_demo  # noqa: E402
from dossier import evidence_snapshot_and_hash, generate_workbook  # noqa: E402
from pilot_store import PilotStore  # noqa: E402

EXPECTED_SHEETS = [
    "Read Me",
    "Experiments",
    "Recommendation Batches",
    "Qualification Gates",
    "Replicate Summary",
    "Calibration Run Summary",
    "Calibration Run Obs",
    "Calibration Form Summary",
    "Calibration Form Obs",
    "Robustness Runs",
    "Approvals",
    "Audit Trail",
]


class WorkbookExportTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.store = PilotStore(Path(self.tempdir.name) / "workspace.db")
        self.project_id = seed_demo(self.store)
        self.user = self.store.authenticate(DEMO_OWNER_EMAIL, DEMO_OWNER_PASSWORD)

    def tearDown(self):
        self.tempdir.cleanup()

    def test_workbook_contains_full_project_as_tabs(self):
        snapshot, evidence_hash = evidence_snapshot_and_hash(self.store, self.project_id)
        self.store.sign_approval(
            self.project_id,
            stage="screening",
            signer_user_id=self.user["id"],
            typed_name=self.user["display_name"],
            password=DEMO_OWNER_PASSWORD,
            signature_meaning="I reviewed the evidence for this stage and approve progression.",
            evidence_hash=evidence_hash,
            evidence_snapshot=snapshot,
        )
        workbook_bytes, manifest = generate_workbook(
            self.store, self.project_id, generated_by_user_id=self.user["id"]
        )
        self.assertEqual(manifest["sheets"], EXPECTED_SHEETS)

        workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), read_only=True)
        self.assertEqual(workbook.sheetnames, EXPECTED_SHEETS)

        experiments_rows = list(workbook["Experiments"].iter_rows(values_only=True))
        self.assertEqual(len(experiments_rows) - 1, 88)  # header + all demo lots

        approvals_rows = list(workbook["Approvals"].iter_rows(values_only=True))
        approvals_header = [str(value) for value in approvals_rows[0]]
        self.assertIn("evidence_hash", approvals_header)
        self.assertNotIn("evidence_snapshot", approvals_header)
        self.assertEqual(len(approvals_rows) - 1, 1)

        cover_values = [
            str(value)
            for row in workbook["Read Me"].iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        self.assertIn(manifest["scientific_evidence_sha256"], cover_values)
        workbook.close()

    def test_workbook_export_is_audited(self):
        generate_workbook(self.store, self.project_id, generated_by_user_id=self.user["id"])
        audit = self.store.audit_log(self.project_id)
        self.assertIn("workbook_exported", set(audit["event_type"]))


if __name__ == "__main__":
    unittest.main()
